#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Feb 14 22:30:52 2019

@author: pandit
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Feb 14 21:29:30 2019

@author: pandit
"""
#Reading an excel file using Python
import xlrd

#Give the location of the file
loc=("/home/pandit/Downloads/Projects/2019-excel-calendar-holidays-landscape.xls")

#To open workbook
wb=xlrd.open_workbook(loc)
sheet=wb.sheet_by_index(1)

#For row 0 and column 0
sheet.cell_value(0,0)

#Extracting number of rows
print(sheet.nrows)

print(sheet.ncols)

for i in range(1,sheet.ncols):
    print(sheet.cell_value(0,i))
    
for i in range(sheet.nrows):
    print(sheet.cell_value(i,0))
    
#Extract particular row value
sheet.cell_value(0,0)

print(sheet.row_values(i))

#will print a particular row value
import openpyxl
path=("/home/pandit/Downloads/Projects/2019-excel-calendar-holidays-landscape.xls")

#workbook object is created
wb_obj=openpyxl.load_workbook(path)

sheet_obj=wb_obj.active

max_col=sheet_obj.max_column

#will print particular row values
for i in range(1,max_col+1):
    cell_obj=sheet_obj.cell(row=2,column=i)
    print(cell_obj.value,end=" ")

    