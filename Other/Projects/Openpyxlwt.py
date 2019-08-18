#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Feb 15 08:27:03 2019

@author: pandit
"""

#import openpyxl  module
import  openpyxl

#call a workbook() function of openpyxl
#to create a new blank Workbook object

wb=openpyxl.Workbook()

#Get workbook active sheet from the active attribute
sheet=wb.active

#Once have the Worksheet object,one can get its name from the title attribute
sheet_title=sheet.title

print("Active sheet title:"+sheet_title)



#Code #2 Program to change the Title name
#import openpyxl module
import openpyxl

#call a workbook() function of openpyxl
#to create a new blank Workbook object

wb=openpyxl.Workbook()

#Get workbook active sheet from the active attribute
sheet=wb.active

#One can change the name of the titile
sheet.title="sheet1"

print("sheet name is renamed as:"+sheet.title)

#3 Program to write to an Excel sheet

#import openpyxl module
import openpyxl

#call a workbook() function of openpyxl to create a new
#blank workbook  object
wb=openpyxl.Workbook()

#Get workbook  active sheet from the active attribute
sheet=wb.active

"""
cell objects also have row,column
and coordinate attributes that provide 
location inormation for the cell.
"""
c1=sheet.cell(row=1,column=1)

#writing values to cells
c1.value="ANKIT"

c2=sheet.cell(row=1,column=2)
c2.value="RAI"

'''
Once have a worksheet object,one can access
a cell object by its name also. A2 means 
column=1 and row=2
'''

c3=sheet['A2']
c3.value="RAHUL"

#B2 means column=2 & row=2
c4=sheet['B2']
c4.value="RAI"


##4 :Program to add Sheets in the Workbook

#import openpyxl module
import openpyxl

#call a workbook() function of openpyxl
#to create a new blank workbook object

wb=openpyxl.Workbook()

sheet=wb.active

#Sheets can be added to workbook with the
#workbook  objects create sheet() method.
wb.create_sheet(index=1,title="demo sheet1")

wb.save("/home/pandit/Downloads/Projects/demo.xlsx")