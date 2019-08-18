#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Feb 15 08:52:47 2019

@author: pandit
"""

#Code #1 : Program to set the dimensions of the cells.

#import openpyxl  module
import  openpyxl

#Call a Workbook() function of openpyxl to create a
#new blank Workbook object

wb=openpyxl.Workbook()

#Get workbook  active sheet from
#the active attribute.
sheet=wb.active

#Writing to the specified cell
sheet.cell(row=1,column=1).value='hello'

sheet.cell(row=2,column=2).value='everyone'

#set the height of the row
sheet.row_dimensions[1].height=70

#set the width of the column
sheet.column_dimensions['B'].width=20

#save the file
wb.save("/home/pandit/Downloads/Projects/dimensions.xlsx")