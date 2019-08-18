#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Feb 15 07:39:38 2019

@author: pandit
"""
#import openpyxl module
import openpyxl

#Give the location of the file
path=("/home/pandit/Downloads/Projects/Sample-Sales-Data.xlsx")

#To open the workbook
#workbook object is created
wb_obj=openpyxl.load_workbook(path)

#Get workbook active sheet object
#from the active attribute
sheet_obj=wb_obj.active

#cell object is created by using 
#sheet objects cell() method
cell_obj=sheet_obj.cell(row=1,column=1)

#print value of cell object
#using the value attribute
print(cell_obj.value)

print(sheet_obj.max_row)

#print total number of column
print(sheet_obj.max_column)

max_col=sheet_obj.max_column

#loop will print all columns name
for i in range(1,max_col+1):
    cell_obj=sheet_obj.cell(row=1,column=i)
    print(cell_obj.value)
    
#Print first column value
    
    
#importing openpyxl module
import openpyxl

#Give the location of the file
path=("/home/pandit/Downloads/Projects/Sample-Sales-Data.xlsx")

#workbook object is created
wb_obj=openpyxl.load_workbook(path)

sheet_obj=wb_obj.active
m_row=sheet_obj.max_row

#Loop will print all values of first column
for i in range(1,m_row+1):
       cell_obj=sheet_obj.cell(row=i,column=1)
       print(cell_obj.value)

max_col=sheet_obj.max_column

#will print particular row value
for i in range(1,max_col+1):
    cell_obj=sheet_obj.cell(row=2,column=i)
    print(cell_obj.value,end=" ")